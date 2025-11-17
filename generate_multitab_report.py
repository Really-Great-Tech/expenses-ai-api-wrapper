import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

def load_json_file(filepath):
    """Load JSON file if it exists, return None otherwise"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return None

def load_markdown_file(filepath):
    """Load markdown file if it exists, return None otherwise"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
    return None

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Handle arrays by converting to string representation
            items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)

def create_section_header(ws, row, title, start_col=1, end_col=10):
    """Create a formatted section header"""
    # Merge cells for the header
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    
    # Set the title
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    return row + 1

def add_dataframe_to_worksheet(ws, df, start_row, title=""):
    """Add a dataframe to worksheet with formatting"""
    if title:
        start_row = create_section_header(ws, start_row, title, end_col=len(df.columns))
        start_row += 1

    # Add headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Add data
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Special formatting for LLM VALIDATION - DIMENSION RESULTS issues column
            if title == "LLM VALIDATION - DIMENSION RESULTS":
                # Find the issues column by checking the column name directly
                if 'issues' in df.columns[col_idx-1].lower():
                    # Get the dimension value from the same row (column 2, which is index 1)
                    dimension_value = str(row_data[1]) if len(row_data) > 1 else ""
                    
                    # Apply unique color coding for each dimension
                    if 'factual grounding' in dimension_value.lower():
                        # Light Red for factual grounding
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif 'knowledge base adherence' in dimension_value.lower():
                        # Light Orange for knowledge base adherence
                        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    elif 'compliance accuracy' in dimension_value.lower():
                        # Light Yellow for compliance accuracy
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    elif 'issue categorization' in dimension_value.lower():
                        # Light Blue for issue categorization
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif 'recommendation validity' in dimension_value.lower():
                        # Light Purple for recommendation validity
                        cell.fill = PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid")
                    elif 'hallucination detection' in dimension_value.lower():
                        # Light Pink for hallucination detection
                        cell.fill = PatternFill(start_color="FFE6F3", end_color="FFE6F3", fill_type="solid")
                    elif not value or str(value).strip() == "":
                        # Light Green for no issues
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                    else:
                        # Light Gray for any other dimensions
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    
                    # Set text wrapping for better readability
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            # Special formatting for ISSUES SECTION - add text wrapping and color coding
            if title == "ISSUES SECTION":
                # Apply text wrapping to description (column 3), recommendation (column 4), knowledge_base_reference (column 5), and judge_explanation (column 8) columns
                if col_idx in [3, 4, 5, 8]:  # description, recommendation, knowledge_base_reference, judge_explanation columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
                # Apply color coding to aggregated_issue_validation_score (column 7) and judge_explanation (column 8) based on validation score
                if col_idx in [7, 8]:  # aggregated_issue_validation_score and judge_explanation columns
                    # Get the aggregated validation score from column 7 of the same row
                    validation_score_value = row_data[6] if len(row_data) > 6 else None  # Column 7 is index 6
                    
                    try:
                        score = float(validation_score_value) if validation_score_value else 0
                        
                        # Color coding based on validation score ranges
                        if score >= 90:
                            # Excellent validation (90-100) - Dark Green
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif score >= 80:
                            # Good validation (80-89) - Light Green
                            cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                        elif score >= 70:
                            # Moderate validation (70-79) - Light Yellow
                            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        elif score >= 60:
                            # Fair validation (60-69) - Light Orange
                            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                        elif score >= 50:
                            # Poor validation (50-59) - Light Red
                            cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                        elif score > 0:
                            # Very poor validation (1-49) - Red
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        else:
                            # No score or zero - Light Gray
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    except (ValueError, TypeError):
                        # If score is not a valid number, use light gray
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Special formatting for markdown content
            if title == "MARKDOWN SECTION":
                if col_idx == 3:  # markdown_content column (now column 3 due to content_part column)
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    # Set column width to be wider for markdown content
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 120
                    # Set row height to accommodate more text (reduced since we're chunking)
                    ws.row_dimensions[row_idx].height = 300
                    # Use a monospace-like font for better readability
                    cell.font = Font(name='Courier New', size=10)
                elif col_idx == 2:  # content_part column
                    # Format the part indicator column
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 20
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            

    return row_idx + 3  # Return next available row with spacing

def create_classification_section(results_data, filename):
    """Create classification section dataframe"""
    if not results_data or 'classification' not in results_data:
        return pd.DataFrame()

    classification = results_data['classification']
    flattened = flatten_dict(classification)

    # Create single row dataframe
    data = {'file': filename}
    data.update(flattened)
    data['QA'] = ''
    data['note'] = ''

    return pd.DataFrame([data])

def create_extraction_section(results_data, filename):
    """Create extraction section dataframe"""
    if not results_data or 'extraction' not in results_data:
        return pd.DataFrame()

    extraction = results_data['extraction']
    rows = []

    for key, value in extraction.items():
        if isinstance(value, dict):
            # Handle nested objects
            for sub_key, sub_value in value.items():
                rows.append({
                    'file_name': filename,
                    'field_key': f"{key}_{sub_key}",
                    'field_value': str(sub_value) if sub_value is not None else '',
                    'QA': '',
                    'note': ''
                })
        elif isinstance(value, list):
            # Handle arrays
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    for sub_key, sub_value in item.items():
                        rows.append({
                            'file_name': filename,
                            'field_key': f"{key}[{i}]_{sub_key}",
                            'field_value': str(sub_value) if sub_value is not None else '',
                            'QA': '',
                            'note': ''
                        })
                else:
                    rows.append({
                        'file_name': filename,
                        'field_key': f"{key}[{i}]",
                        'field_value': str(item) if item is not None else '',
                        'QA': '',
                        'note': ''
                    })
        else:
            rows.append({
                'file_name': filename,
                'field_key': key,
                'field_value': str(value) if value is not None else '',
                'QA': '',
                'note': ''
            })

    return pd.DataFrame(rows)

def create_issues_section(results_data, filename):
    """Create issues section dataframe"""
    if not results_data or 'compliance' not in results_data:
        return pd.DataFrame()

    compliance = results_data['compliance']
    validation = compliance.get('validation_result', {})
    issues = validation.get('issues', [])

    # Load validation data to get aggregated issue validation scores and judge explanations
    import glob
    validation_data = None
    timestamped_validation_files = glob.glob(f'validation_results/{filename}_validation_*.json')
    if timestamped_validation_files:
        # Use the most recent timestamped validation file
        validation_data = load_json_file(sorted(timestamped_validation_files)[-1])

    rows = []
    for i, issue in enumerate(issues, 1):
        # Initialize default values for new columns
        aggregated_score = ''
        judge_explanations = ''
        
        # Extract aggregated issue validation scores and judge explanations if available
        if validation_data and 'issue_validation_scores' in validation_data:
            issue_validation_scores = validation_data['issue_validation_scores']
            
            # Find the matching issue by index (i-1 since we start from 1)
            matching_issue = None
            for issue_score in issue_validation_scores:
                if issue_score.get('issue_index') == i - 1:
                    matching_issue = issue_score
                    break
            
            if matching_issue:
                aggregated_score = matching_issue.get('overall_validation_score', '')
                
                # Collect judge explanations from dimension results
                explanations = []
                if 'dimension_results' in validation_data:
                    for dimension_result in validation_data['dimension_results']:
                        if 'issue_validation_scores' in dimension_result:
                            for issue_val_score in dimension_result['issue_validation_scores']:
                                if issue_val_score.get('issue_index') == i - 1:
                                    dimension_name = dimension_result.get('dimension', '').replace('_', ' ').title()
                                    explanation = issue_val_score.get('judge_explanation', '')
                                    if explanation:
                                        explanations.append(f"{dimension_name}: {explanation}")
                
                judge_explanations = ' | '.join(explanations)
        
        rows.append({
            'index': i,
            'issue_type': issue.get('issue_type', ''),
            'description': issue.get('description', ''),
            'recommendation': issue.get('recommendation', ''),
            'knowledge_base_reference': issue.get('knowledge_base_reference', ''),
            'confidence_score': issue.get('confidence_score', ''),
            'aggregated_issue_validation_score': aggregated_score,
            'judge_explanation': judge_explanations,
            'QA': '',
            'note': ''
        })

    return pd.DataFrame(rows)

def create_validation_overall_section(validation_data, filename):
    """Create validation overall assessment section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    overall = validation_data['validation_report'].get('overall_assessment', {})
    data = {'file': filename}
    data.update(overall)
    data['QA'] = ''
    data['note'] = ''
    
    return pd.DataFrame([data])

def create_validation_dimensions_section(validation_data, filename):
    """Create validation dimension details section"""
    if not validation_data or 'detailed_analysis' not in validation_data:
        return pd.DataFrame()
    
    dimension_details = validation_data['detailed_analysis'].get('dimension_details', {})
    rows = []
    
    for dimension, details in dimension_details.items():
        row = {
            'file': filename,
            'dimension': dimension,
            'confidence_score': details.get('confidence_score', ''),
            'reliability_level': details.get('reliability_level', ''),
            'summary': details.get('summary', ''),
            'total_issues': details.get('total_issues', ''),
            'QA': '',
            'note': ''
        }
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_validation_summary_section(validation_data, filename):
    """Create dimensional analysis summary section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    dimensional_summary = validation_data['validation_report'].get('dimensional_analysis_summary', {})
    rows = []
    
    for dimension, details in dimensional_summary.items():
        row = {
            'file': filename,
            'dimension': dimension,
            'confidence': details.get('confidence', ''),
            'reliability': details.get('reliability', ''),
            'issues_count': details.get('issues_count', ''),
            'QA': '',
            'note': ''
        }
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_validation_critical_section(validation_data, filename):
    """Create critical issues summary section"""
    if not validation_data or 'validation_report' not in validation_data:
        return pd.DataFrame()
    
    critical_issues = validation_data['validation_report'].get('critical_issues_summary', {}).get('issues', [])
    rows = []
    
    for i, issue in enumerate(critical_issues, 1):
        rows.append({
            'file': filename,
            'index': i,
            'critical_issue': issue,
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_markdown_section(markdown_content, filename):
    """Create markdown section dataframe with chunked content for better readability"""
    if not markdown_content:
        return pd.DataFrame()

    # Clean up the markdown content for better display
    # Remove the metadata header (everything before the first ---)
    lines = markdown_content.split('\n')
    content_start = 0
    for i, line in enumerate(lines):
        if line.strip() == '---' and i > 0:  # Find the closing --- of metadata
            content_start = i + 1
            break

    # Join the actual content (skip metadata)
    clean_content = '\n'.join(lines[content_start:]).strip()

    # Split content into chunks for better display (approximately 50 lines per chunk)
    content_lines = clean_content.split('\n')
    chunk_size = 50
    rows = []

    if len(content_lines) <= chunk_size:
        # If content is short enough, show in single row
        rows.append({
            'file_name': filename,
            'content_part': 'Complete Content',
            'markdown_content': clean_content,
            'QA': '',
            'note': ''
        })
    else:
        # Split into multiple chunks for very long content
        for i in range(0, len(content_lines), chunk_size):
            chunk_lines = content_lines[i:i + chunk_size]
            chunk_content = '\n'.join(chunk_lines)
            part_num = (i // chunk_size) + 1
            total_parts = (len(content_lines) + chunk_size - 1) // chunk_size

            rows.append({
                'file_name': filename,
                'content_part': f'Part {part_num} of {total_parts}',
                'markdown_content': chunk_content,
                'QA': '',
                'note': ''
            })

    return pd.DataFrame(rows)

def create_timing_section(results_data, filename):
    """Create timing section dataframe"""
    if not results_data or 'timing' not in results_data:
        return pd.DataFrame()

    timing = results_data['timing']
    rows = []

    # Total processing time
    total_time = timing.get('total_processing_time_seconds', 'N/A')
    rows.append({
        'file_name': filename,
        'timing_type': 'Total Processing',
        'phase_or_metric': 'Total Processing Time',
        'value': f"{total_time} seconds",
        'details': 'Complete end-to-end processing time',
        'QA': '',
        'note': ''
    })

    # Performance metrics
    if 'performance_metrics' in timing:
        perf_metrics = timing['performance_metrics']
        for metric, value in perf_metrics.items():
            metric_name = metric.replace('_', ' ').title()
            rows.append({
                'file_name': filename,
                'timing_type': 'Performance Metric',
                'phase_or_metric': metric_name,
                'value': f"{value} seconds" if 'seconds' in metric else str(value),
                'details': '',
                'QA': '',
                'note': ''
            })

    # Agent performance details (individual phases)
    if 'agent_performance' in timing:
        agent_performance = timing['agent_performance']
        for phase, details in agent_performance.items():
            if isinstance(details, dict):
                duration = details.get('duration_seconds', 'N/A')
                model_used = details.get('model_used', details.get('document_reader_used', ''))
                execution_mode = details.get('execution_mode', '')

                detail_str = f"Model: {model_used}" if model_used else ""
                if execution_mode:
                    detail_str += f", Mode: {execution_mode}" if detail_str else f"Mode: {execution_mode}"

                rows.append({
                    'file_name': filename,
                    'timing_type': 'Agent Performance',
                    'phase_or_metric': phase.replace('_', ' ').title(),
                    'value': f"{duration} seconds",
                    'details': detail_str,
                    'QA': '',
                    'note': ''
                })

    return pd.DataFrame(rows)

def create_image_quality_overall_section(quality_data, filename):
    """Create image quality overall assessment section"""
    if not quality_data:
        return pd.DataFrame()
    
    overall = quality_data.get('overall_assessment', {})
    image_type = quality_data.get('image_type_detection', {})
    
    data = {
        'file': filename,
        'overall_score': overall.get('score', ''),
        'quality_level': overall.get('level', ''),
        'quality_passed': overall.get('pass_fail', ''),
        'processing_time_seconds': quality_data.get('processing_time_seconds', ''),
        'image_type': image_type.get('image_subtype', ''),
        'is_digital_screenshot': image_type.get('is_digital_screenshot', ''),
        'confidence': image_type.get('confidence', ''),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_image_quality_detailed_section(quality_data, filename):
    """Create image quality detailed metrics section"""
    if not quality_data or 'detailed_results' not in quality_data:
        return pd.DataFrame()
    
    detailed = quality_data['detailed_results']
    rows = []
    
    # Resolution metrics
    if 'resolution' in detailed:
        res = detailed['resolution']
        dimensions = res.get('dimensions', {})
        dpi = res.get('dpi', {})
        quality = res.get('quality', {})
        
        key_metrics = f"Dimensions: {dimensions.get('width', '')}x{dimensions.get('height', '')}, DPI: {dpi.get('average', '')}, Megapixels: {dimensions.get('megapixels', '')}"
        recommendations = '; '.join(res.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Resolution',
            'score': quality.get('score', ''),
            'level': quality.get('level', ''),
            'meets_requirements': quality.get('meets_ocr_requirements', ''),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Blur metrics
    if 'blur' in detailed:
        blur = detailed['blur']
        metrics = blur.get('metrics', {})
        motion_blur = blur.get('motion_blur', {})
        
        key_metrics = f"Laplacian Variance: {metrics.get('laplacian_variance', '')}, Motion Blur: {motion_blur.get('detected', '')}, Direction: {motion_blur.get('direction', '')}"
        recommendations = '; '.join(blur.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Blur',
            'score': metrics.get('blur_score', ''),
            'level': metrics.get('blur_level', ''),
            'meets_requirements': not metrics.get('is_blurry', True),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Glare metrics
    if 'glare' in detailed:
        glare = detailed['glare']
        exposure = glare.get('exposure_metrics', {})
        analysis = glare.get('glare_analysis', {})
        
        key_metrics = f"Mean Brightness: {exposure.get('mean_brightness', '')}, Overexposed: {exposure.get('overexposed_percent', '')}%, Glare Spots: {analysis.get('num_glare_spots', '')}"
        recommendations = '; '.join(glare.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Glare',
            'score': analysis.get('glare_score', ''),
            'level': analysis.get('glare_level', ''),
            'meets_requirements': not exposure.get('is_overexposed', False),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Completeness metrics
    if 'completeness' in detailed:
        completeness = detailed['completeness']
        edge_analysis = completeness.get('edge_analysis', {})
        corner_analysis = completeness.get('corner_analysis', {})
        
        key_metrics = f"Boundary Detected: {completeness.get('boundary_detected', '')}, Edge Coverage: {edge_analysis.get('edge_coverage', '')}%, Visible Corners: {corner_analysis.get('visible_corners', '')}"
        recommendations = '; '.join(completeness.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Completeness',
            'score': completeness.get('completeness_score', ''),
            'level': completeness.get('completeness_level', ''),
            'meets_requirements': completeness.get('boundary_detected', ''),
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    # Damage metrics
    if 'damage' in detailed:
        damage = detailed['damage']
        stain = damage.get('stain_analysis', {})
        tear = damage.get('tear_analysis', {})
        fold = damage.get('fold_analysis', {})
        
        key_metrics = f"Damage Types: {', '.join(damage.get('damage_types', []))}, Stains: {stain.get('count', 0)}, Tears: {tear.get('count', 0)}, Folds: {fold.get('count', 0)}"
        recommendations = '; '.join(damage.get('recommendations', []))
        
        rows.append({
            'file': filename,
            'quality_aspect': 'Damage',
            'score': damage.get('damage_score', ''),
            'level': damage.get('damage_level', ''),
            'meets_requirements': len(damage.get('damage_types', [])) == 0,
            'key_metrics': key_metrics,
            'recommendations': recommendations,
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_image_quality_breakdown_section(quality_data, filename):
    """Create image quality score breakdown section"""
    if not quality_data or 'score_breakdown' not in quality_data:
        return pd.DataFrame()
    
    breakdown = quality_data['score_breakdown']
    rows = []
    
    for component, details in breakdown.items():
        rows.append({
            'file': filename,
            'quality_component': component.title(),
            'individual_score': details.get('score', ''),
            'weight': details.get('weight', ''),
            'contribution': details.get('contribution', ''),
            'QA': '',
            'note': ''
        })
    
    return pd.DataFrame(rows)

def create_citation_section(citation_data, filename):
    """Create citation section dataframe with field-level citations"""
    if not citation_data or 'citations' not in citation_data:
        return pd.DataFrame()

    # Handle nested citations structure: citations.citations
    citations_section = citation_data['citations']
    if 'citations' in citations_section:
        citations = citations_section['citations']
    else:
        citations = citations_section
    rows = []
    
    for field_name, citation_info in citations.items():
        field_citation = citation_info.get('field_citation', {})
        value_citation = citation_info.get('value_citation', {})
        
        # Handle case where value_citation might be None
        if value_citation is None:
            value_citation = {}
        
        row = {
            'file_name': filename,
            'field_name': field_name,
            'field_source_text': field_citation.get('source_text', ''),
            'field_confidence': field_citation.get('confidence', ''),
            'field_source_location': field_citation.get('source_location', ''),
            'field_match_type': field_citation.get('match_type', ''),
            'value_source_text': value_citation.get('source_text', ''),
            'value_confidence': value_citation.get('confidence', ''),
            'value_source_location': value_citation.get('source_location', ''),
            'value_match_type': value_citation.get('match_type', ''),
            'context': field_citation.get('context', ''),
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_citation_metadata_section(citation_data, filename):
    """Create citation metadata summary section"""
    if not citation_data or 'citations' not in citation_data:
        return pd.DataFrame()

    # Handle nested citations structure: citations.metadata
    citations_section = citation_data['citations']
    if 'metadata' not in citations_section:
        return pd.DataFrame()

    metadata = citations_section['metadata']
    
    data = {
        'file_name': filename,
        'total_fields_analyzed': metadata.get('total_fields_analyzed', 0),
        'fields_with_field_citations': metadata.get('fields_with_field_citations', 0),
        'fields_with_value_citations': metadata.get('fields_with_value_citations', 0),
        'average_confidence': metadata.get('average_confidence', 0),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_llm_quality_overall_section(llm_quality_data, filename):
    """Create LLM quality overall assessment section"""
    if not llm_quality_data:
        return pd.DataFrame()

    data = {
        'file_name': filename,
        'assessment_method': llm_quality_data.get('assessment_method', ''),
        'model_used': llm_quality_data.get('model_used', ''),
        'overall_quality_score': llm_quality_data.get('overall_quality_score', ''),
        'quality_score': llm_quality_data.get('quality_score', ''),
        'quality_level': llm_quality_data.get('quality_level', ''),
        'suitable_for_extraction': llm_quality_data.get('suitable_for_extraction', ''),
        'timestamp': llm_quality_data.get('timestamp', ''),
        'QA': '',
        'note': ''
    }

    return pd.DataFrame([data])

def create_llm_quality_detailed_section(llm_quality_data, filename):
    """Create LLM quality detailed assessment section"""
    if not llm_quality_data:
        return pd.DataFrame()
    
    # Define the quality aspects to extract
    quality_aspects = [
        'blur_detection',
        'contrast_assessment',
        'glare_identification',
        'water_stains',
        'tears_or_folds',
        'cut_off_detection',
        'missing_sections',
        'obstructions'
    ]
    
    rows = []
    
    for aspect in quality_aspects:
        if aspect in llm_quality_data:
            aspect_data = llm_quality_data[aspect]
            
            row = {
                'file_name': filename,
                'quality_aspect': aspect.replace('_', ' ').title(),
                'detected': aspect_data.get('detected', ''),
                'severity_level': aspect_data.get('severity_level', ''),
                'confidence_score': aspect_data.get('confidence_score', ''),
                'quantitative_measure': aspect_data.get('quantitative_measure', ''),
                'description': aspect_data.get('description', ''),
                'recommendation': aspect_data.get('recommendation', ''),
                'QA': '',
                'note': ''
            }
            
            rows.append(row)
    
    return pd.DataFrame(rows)

def create_llm_validation_overall_section(llm_validation_data, filename):
    """Create LLM validation overall assessment section"""
    if not llm_validation_data:
        return pd.DataFrame()
    
    data = {
        'file_name': filename,
        'overall_score': round(llm_validation_data.get('overall_score', 0) * 100, 1),  # Convert to percentage
        'dimensions_count': llm_validation_data.get('dimensions_count', 0),
        'overall_reliability': llm_validation_data.get('overall_reliability', ''),
        'judge_models': ', '.join(llm_validation_data.get('metadata', {}).get('judge_models', [])),
        'total_processing_time_seconds': llm_validation_data.get('metadata', {}).get('timing', {}).get('total_validation_time_seconds', ''),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_llm_validation_dimensions_section(llm_validation_data, filename):
    """Create LLM validation dimension results section"""
    if not llm_validation_data or 'dimension_results' not in llm_validation_data:
        return pd.DataFrame()
    
    dimension_results = llm_validation_data['dimension_results']
    rows = []
    
    for result in dimension_results:
        # Convert issues list to string for display
        issues_text = '; '.join(result.get('issues', []))
        
        row = {
            'file_name': filename,
            'dimension': result.get('dimension', '').replace('_', ' ').title(),
            'confidence_score': round(result.get('confidence_score', 0) * 100, 1),  # Convert to percentage
            'reliability_level': result.get('reliability_level', ''),
            'issues': issues_text,
            'judge_models': ', '.join(result.get('judge_models', [])),
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_llm_validation_judge_details_section(llm_validation_data, filename):
    """Create LLM validation judge details section"""
    if not llm_validation_data or 'dimension_results' not in llm_validation_data:
        return pd.DataFrame()
    
    dimension_results = llm_validation_data['dimension_results']
    rows = []
    
    for result in dimension_results:
        dimension = result.get('dimension', '').replace('_', ' ').title()
        judge_details = result.get('judge_details', [])
        
        for judge in judge_details:
            row = {
                'file_name': filename,
                'dimension': dimension,
                'judge_model': judge.get('model_name', ''),
                'confidence_score': round(judge.get('confidence_score', 0) * 100, 1),  # Convert to percentage
                'response_summary': judge.get('response', '')[:200] + '...' if len(judge.get('response', '')) > 200 else judge.get('response', ''),
                'QA': '',
                'note': ''
            }
            
            rows.append(row)
    
    return pd.DataFrame(rows)

def create_llm_validation_recommendations_section(llm_validation_data, filename):
    """Create LLM validation recommendations section"""
    if not llm_validation_data or 'recommendations' not in llm_validation_data:
        return pd.DataFrame()
    
    recommendations = llm_validation_data['recommendations']
    rows = []
    
    for i, recommendation in enumerate(recommendations, 1):
        row = {
            'file_name': filename,
            'recommendation_index': i,
            'recommendation': recommendation,
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_llm_validation_critical_issues_section(llm_validation_data, filename):
    """Create LLM validation critical issues section"""
    if not llm_validation_data or 'critical_issues' not in llm_validation_data:
        return pd.DataFrame()
    
    critical_issues = llm_validation_data['critical_issues']
    if not critical_issues:  # Return empty DataFrame if no critical issues
        return pd.DataFrame()
    
    rows = []
    
    for i, issue in enumerate(critical_issues, 1):
        row = {
            'file_name': filename,
            'critical_issue_index': i,
            'critical_issue': issue,
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_new_validation_overall_section(validation_data, filename):
    """Create new validation overall assessment section from timestamped validation files"""
    if not validation_data:
        return pd.DataFrame()
    
    data = {
        'file_name': filename,
        'overall_score': round(validation_data.get('overall_score', 0) * 100, 1),  # Convert to percentage
        'dimensions_count': validation_data.get('dimensions_count', 0),
        'overall_reliability': validation_data.get('overall_reliability', ''),
        'critical_issues_count': len(validation_data.get('critical_issues', [])),
        'judge_models': ', '.join(validation_data.get('metadata', {}).get('judge_models', [])),
        'total_processing_time_seconds': validation_data.get('metadata', {}).get('timing', {}).get('total_validation_time_seconds', ''),
        'validation_timestamp': validation_data.get('timestamp', ''),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def create_new_validation_dimensions_section(validation_data, filename):
    """Create new validation dimension results section (excluding summary and judge details)"""
    if not validation_data or 'dimension_results' not in validation_data:
        return pd.DataFrame()
    
    dimension_results = validation_data['dimension_results']
    rows = []
    
    for result in dimension_results:
        # Convert issues list to string for display
        issues_text = '; '.join(result.get('issues', []))
        
        row = {
            'file_name': filename,
            'dimension': result.get('dimension', '').replace('_', ' ').title(),
            'confidence_score': round(result.get('confidence_score', 0) * 100, 1),  # Convert to percentage
            'reliability_level': result.get('reliability_level', ''),
            'issues_count': len(result.get('issues', [])),
            'issues': issues_text,
            # Explicitly exclude 'summary' as requested
            # Explicitly exclude 'judge_details' as requested
            'judge_models': ', '.join(result.get('judge_models', [])),
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_new_validation_recommendations_section(validation_data, filename):
    """Create new validation recommendations section"""
    if not validation_data or 'recommendations' not in validation_data:
        return pd.DataFrame()
    
    recommendations = validation_data['recommendations']
    rows = []
    
    for i, recommendation in enumerate(recommendations, 1):
        row = {
            'file_name': filename,
            'recommendation_index': i,
            'recommendation': recommendation,
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_new_validation_critical_issues_section(validation_data, filename):
    """Create new validation critical issues section"""
    if not validation_data or 'critical_issues' not in validation_data:
        return pd.DataFrame()
    
    critical_issues = validation_data['critical_issues']
    if not critical_issues:  # Return empty DataFrame if no critical issues
        return pd.DataFrame()
    
    rows = []
    
    for i, issue in enumerate(critical_issues, 1):
        row = {
            'file_name': filename,
            'critical_issue_index': i,
            'critical_issue': issue,
            'QA': '',
            'note': ''
        }
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_new_validation_metadata_section(validation_data, filename):
    """Create new validation metadata section"""
    if not validation_data or 'metadata' not in validation_data:
        return pd.DataFrame()
    
    metadata = validation_data['metadata']
    timing = metadata.get('timing', {})
    
    data = {
        'file_name': filename,
        'validation_version': metadata.get('validation_version', ''),
        'judge_models_count': len(metadata.get('judge_models', [])),
        'judge_models': ', '.join(metadata.get('judge_models', [])),
        'processing_time_ms': metadata.get('processing_time_ms', ''),
        'total_validation_time_seconds': timing.get('total_validation_time_seconds', ''),
        'validation_start_time': timing.get('validation_start_time', ''),
        'validation_end_time': timing.get('validation_end_time', ''),
        'dimensions_validated': metadata.get('context', {}).get('dimensions_validated', ''),
        'judge_panel_size': metadata.get('context', {}).get('judge_panel_size', ''),
        'QA': '',
        'note': ''
    }
    
    return pd.DataFrame([data])

def auto_adjust_column_widths(ws):
    """Auto-adjust column widths for a worksheet"""
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_worksheet_for_file(wb, filename):
    """Create a worksheet for a single file with all sections"""
    print(f"Processing {filename}...")

    # Load data from all sources with current structure
    results_data = load_json_file(f'results/{filename}_result.json')

    # Try different validation file patterns (keeping fallback for compatibility)
    validation_data = (load_json_file(f'validation_results/{filename}_compliance_validation.json') or
                      load_json_file(f'validation_results/{filename}_validation.json'))

    # Try different LLM validation file patterns
    llm_validation_data = None
    validation_files = [
        f'validation_results/{filename}_llm_validation.json',
        f'validation_results/{filename}_validation_*.json'
    ]
    for pattern in validation_files:
        if '*' in pattern:
            # Handle wildcard pattern for timestamped files
            import glob
            matching_files = glob.glob(pattern.replace('*', '*'))
            if matching_files:
                # Use the most recent file
                llm_validation_data = load_json_file(sorted(matching_files)[-1])
                break
        else:
            llm_validation_data = load_json_file(pattern)
            if llm_validation_data:
                break

    # Load new validation results from timestamped files
    new_validation_data = None
    import glob
    timestamped_validation_files = glob.glob(f'validation_results/{filename}_validation_*.json')
    if timestamped_validation_files:
        # Use the most recent timestamped validation file
        new_validation_data = load_json_file(sorted(timestamped_validation_files)[-1])
        print(f"Found timestamped validation file for {filename}: {sorted(timestamped_validation_files)[-1]}")

    # Try different quality file patterns (keeping fallback for compatibility)
    quality_data = (load_json_file(f'quality_reports/{filename}_page1_quality.json') or
                   load_json_file(f'quality_reports/{filename}_quality.json'))

    # Load markdown from new markdown_extractions directory with textract or zerox suffix
    markdown_content = (load_markdown_file(f'markdown_extractions/{filename}_textract.md') or
                       load_markdown_file(f'markdown_extractions/{filename}_zerox.md') or
                       load_markdown_file(f'markdown_extractions/{filename}.md') or
                       load_markdown_file(f'llamaparse_output/{filename}.md'))  # fallback

    # Citations are now part of the main results file
    citation_data = results_data if results_data and 'citations' in results_data else None

    # Try different LLM quality file patterns (keeping fallback for compatibility)
    llm_quality_data = (load_json_file(f'llm_quality_reports/llm_quality_{filename}_pdf.json') or
                       load_json_file(f'llm_quality_reports/llm_quality_{filename}.json'))

    # Use image_quality_assessment from results if available, otherwise use separate LLM quality file
    if results_data and 'image_quality_assessment' in results_data:
        llm_quality_data = results_data['image_quality_assessment']

    # Create worksheet
    ws = wb.create_sheet(title=filename)
    current_row = 1
    
    # 1. Classification Section
    classification_df = create_classification_section(results_data, filename)
    if not classification_df.empty:
        current_row = add_dataframe_to_worksheet(ws, classification_df, current_row, "CLASSIFICATION SECTION")
    
    # 2. Extraction Results Section
    extraction_df = create_extraction_section(results_data, filename)
    if not extraction_df.empty:
        current_row = add_dataframe_to_worksheet(ws, extraction_df, current_row, "EXTRACTION RESULTS SECTION")
    
    # 3. Issues Section
    issues_df = create_issues_section(results_data, filename)
    if not issues_df.empty:
        current_row = add_dataframe_to_worksheet(ws, issues_df, current_row, "ISSUES SECTION")
    
    # 4. Validation - Overall Assessment
    validation_overall_df = create_validation_overall_section(validation_data, filename)
    if not validation_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_overall_df, current_row, "VALIDATION - OVERALL ASSESSMENT")
    
    # 5. Validation - Dimension Details
    validation_dimensions_df = create_validation_dimensions_section(validation_data, filename)
    if not validation_dimensions_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_dimensions_df, current_row, "VALIDATION - DIMENSION DETAILS")
    
    # 6. Validation - Dimensional Analysis Summary
    validation_summary_df = create_validation_summary_section(validation_data, filename)
    if not validation_summary_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_summary_df, current_row, "VALIDATION - DIMENSIONAL ANALYSIS SUMMARY")
    
    # 7. Validation - Critical Issues
    validation_critical_df = create_validation_critical_section(validation_data, filename)
    if not validation_critical_df.empty:
        current_row = add_dataframe_to_worksheet(ws, validation_critical_df, current_row, "VALIDATION - CRITICAL ISSUES")
    
    # 8. Markdown Section
    markdown_df = create_markdown_section(markdown_content, filename)
    if not markdown_df.empty:
        current_row = add_dataframe_to_worksheet(ws, markdown_df, current_row, "MARKDOWN SECTION")
    
    # 9. Citation Section - Field-Level Citations
    citation_df = create_citation_section(citation_data, filename)
    if not citation_df.empty:
        current_row = add_dataframe_to_worksheet(ws, citation_df, current_row, "CITATION SECTION - FIELD-LEVEL CITATIONS")
    
    # 10. Citation Section - Metadata Summary
    citation_metadata_df = create_citation_metadata_section(citation_data, filename)
    if not citation_metadata_df.empty:
        current_row = add_dataframe_to_worksheet(ws, citation_metadata_df, current_row, "CITATION SECTION - METADATA SUMMARY")
    
    # 11. Image Quality - Overall Assessment
    quality_overall_df = create_image_quality_overall_section(quality_data, filename)
    if not quality_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, quality_overall_df, current_row, "IMAGE QUALITY - OVERALL ASSESSMENT")
    
    # 12. Image Quality - Detailed Metrics
    quality_detailed_df = create_image_quality_detailed_section(quality_data, filename)
    if not quality_detailed_df.empty:
        current_row = add_dataframe_to_worksheet(ws, quality_detailed_df, current_row, "IMAGE QUALITY - DETAILED METRICS")
    
    # 13. Timing Section
    timing_df = create_timing_section(results_data, filename)
    if not timing_df.empty:
        current_row = add_dataframe_to_worksheet(ws, timing_df, current_row, "TIMING SECTION")

    # 14. Image Quality - Score Breakdown
    quality_breakdown_df = create_image_quality_breakdown_section(quality_data, filename)
    if not quality_breakdown_df.empty:
        current_row = add_dataframe_to_worksheet(ws, quality_breakdown_df, current_row, "IMAGE QUALITY - SCORE BREAKDOWN")
    
    # 14. LLM Quality - Overall Assessment
    llm_quality_overall_df = create_llm_quality_overall_section(llm_quality_data, filename)
    if not llm_quality_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_quality_overall_df, current_row, "LLM QUALITY - OVERALL ASSESSMENT")
    
    # 15. LLM Quality - Detailed Assessment
    llm_quality_detailed_df = create_llm_quality_detailed_section(llm_quality_data, filename)
    if not llm_quality_detailed_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_quality_detailed_df, current_row, "LLM QUALITY - DETAILED ASSESSMENT")
    
    # 16. LLM Validation - Overall Assessment
    llm_validation_overall_df = create_llm_validation_overall_section(llm_validation_data, filename)
    if not llm_validation_overall_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_validation_overall_df, current_row, "LLM VALIDATION - OVERALL ASSESSMENT")
    
    # 17. LLM Validation - Dimension Results
    llm_validation_dimensions_df = create_llm_validation_dimensions_section(llm_validation_data, filename)
    if not llm_validation_dimensions_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_validation_dimensions_df, current_row, "LLM VALIDATION - DIMENSION RESULTS")
    
    # 18. LLM Validation - Recommendations
    llm_validation_recommendations_df = create_llm_validation_recommendations_section(llm_validation_data, filename)
    if not llm_validation_recommendations_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_validation_recommendations_df, current_row, "LLM VALIDATION - RECOMMENDATIONS")
    
    # 19. LLM Validation - Critical Issues (if any)
    llm_validation_critical_df = create_llm_validation_critical_issues_section(llm_validation_data, filename)
    if not llm_validation_critical_df.empty:
        current_row = add_dataframe_to_worksheet(ws, llm_validation_critical_df, current_row, "LLM VALIDATION - CRITICAL ISSUES")
    
    
    # Auto-adjust column widths
    auto_adjust_column_widths(ws)
    
    print(f"Added worksheet: {filename}")

def get_file_list():
    """Get list of files that have corresponding result files"""
    files = set()

    # Primary source: files in results directory (these are the files we actually processed)
    results_dir = 'results'
    if os.path.exists(results_dir):
        for filename in os.listdir(results_dir):
            if filename.endswith('_result.json'):
                # Extract base name by removing _result.json suffix
                base_name = filename.replace('_result.json', '')
                files.add(base_name)

    # Only return files that have actual results
    return sorted(list(files))

def extract_results_data_summary(data):
    """Extract data from results folder JSON for summary"""
    if not data:
        return {}

    # Extract classification data
    classification = data.get('classification', {})

    # Extract compliance issues count
    compliance = data.get('compliance', {})
    validation = compliance.get('validation_result', {})
    issues_count = validation.get('issues_count', 0)

    return {
        'issues_count': issues_count,
        'is_expense': classification.get('is_expense', False),
        'language': classification.get('language', ''),
        'language_confidence': classification.get('language_confidence', 0),
        'classification_confidence': classification.get('classification_confidence', 0)
    }

def extract_validation_data_summary(data):
    """Extract data from validation_results folder JSON for summary"""
    if not data:
        return {}
    
    validation_report = data.get('validation_report', {})
    overall = validation_report.get('overall_assessment', {})
    dimensional = validation_report.get('dimensional_analysis_summary', {})
    
    return {
        'confidence_score': overall.get('confidence_score', 0),
        'is_reliable': overall.get('is_reliable', False),
        'hallucination_detection': dimensional.get('hallucination_detection', {}).get('confidence', 0),
        'compliance_accuracy': dimensional.get('compliance_accuracy', {}).get('confidence', 0),
        'factual_grounding': dimensional.get('factual_grounding', {}).get('confidence', 0),
        'knowledge_base_adherence': dimensional.get('knowledge_base_adherence', {}).get('confidence', 0)
    }

def extract_quality_data_summary(data):
    """Extract data from quality_reports folder JSON for summary"""
    if not data:
        return {}
    
    overall = data.get('overall_assessment', {})
    detailed = data.get('detailed_results', {})
    
    return {
        'quality_score': overall.get('score', 0),
        'quality_level': overall.get('level', ''),
        'quality_passed': overall.get('pass_fail', False),
        'resolution_score': detailed.get('resolution', {}).get('quality', {}).get('score', 0),
        'blur_score': detailed.get('blur', {}).get('metrics', {}).get('blur_score', 0),
        'glare_score': detailed.get('glare', {}).get('glare_analysis', {}).get('glare_score', 0),
        'completeness_score': detailed.get('completeness', {}).get('completeness_score', 0),
        'damage_score': detailed.get('damage', {}).get('damage_score', 0)
    }

def extract_llm_quality_data_summary(data, results_data=None):
    """Extract data from llm_quality_reports folder JSON or image_quality_assessment for summary"""
    # Try to get LLM quality data from image_quality_assessment in results first
    llm_quality_source = None
    if results_data and 'image_quality_assessment' in results_data:
        llm_quality_source = results_data['image_quality_assessment']
    elif data:
        llm_quality_source = data

    if not llm_quality_source:
        return {}

    return {
        'llm_suitable_for_extraction': llm_quality_source.get('suitable_for_extraction', False),
        'llm_model_used': llm_quality_source.get('model_used', ''),
        'llm_blur_detected': llm_quality_source.get('blur_detection', {}).get('detected', False),
        'llm_blur_severity': llm_quality_source.get('blur_detection', {}).get('severity_level', ''),
        'llm_glare_detected': llm_quality_source.get('glare_identification', {}).get('detected', False),
        'llm_glare_severity': llm_quality_source.get('glare_identification', {}).get('severity_level', ''),
        'llm_cut_off_detected': llm_quality_source.get('cut_off_detection', {}).get('detected', False),
        'llm_missing_sections': llm_quality_source.get('missing_sections', {}).get('detected', False)
    }

def extract_business_data(file_base):
    """Extract business-relevant data for a single file"""

    # Initialize row data
    row_data = {'filename': file_base}

    # Load data from current project structure
    results_data = load_json_file(f'results/{file_base}_result.json')

    # Try different validation file patterns (keeping fallback for compatibility)
    validation_data = (load_json_file(f'validation_results/{file_base}_compliance_validation.json') or
                      load_json_file(f'validation_results/{file_base}_validation.json'))

    # Try different quality file patterns (keeping fallback for compatibility)
    quality_data = (load_json_file(f'quality_reports/{file_base}_page1_quality.json') or
                   load_json_file(f'quality_reports/{file_base}_quality.json'))

    # Citations are now part of the main results file
    citation_data = results_data if results_data and 'citations' in results_data else None

    # Try different LLM quality file patterns (keeping fallback for compatibility)
    llm_quality_data = (load_json_file(f'llm_quality_reports/llm_quality_{file_base}_pdf.json') or
                       load_json_file(f'llm_quality_reports/llm_quality_{file_base}.json'))
    
    # Extract language and isExpense from results
    if results_data and 'classification' in results_data:
        classification = results_data['classification']
        row_data['language'] = classification.get('language', 'Unknown')
        row_data['isExpense'] = classification.get('is_expense', False)
    else:
        row_data['language'] = 'Unknown'
        row_data['isExpense'] = False

    # Extract document location from classification results
    if results_data and 'classification' in results_data:
        classification = results_data['classification']
        row_data['document_location'] = classification.get('document_location', 'Unknown')
    else:
        row_data['document_location'] = 'Unknown'

    # Extract issue count from results
    if results_data and 'compliance' in results_data:
        compliance = results_data['compliance']
        validation = compliance.get('validation_result', {})
        row_data['issue_count'] = validation.get('issues_count', 0)
    else:
        row_data['issue_count'] = 0
    
    # Extract validation confidence score from timestamped validation files and normalize to percentage (0-100)
    # First try to load timestamped validation files which have the overall_score at root level
    timestamped_validation_data = None
    import glob
    timestamped_validation_files = glob.glob(f'validation_results/{file_base}_validation_*.json')
    if timestamped_validation_files:
        # Use the most recent timestamped validation file
        timestamped_validation_data = load_json_file(sorted(timestamped_validation_files)[-1])
    
    if timestamped_validation_data and 'overall_score' in timestamped_validation_data:
        validation_score = timestamped_validation_data.get('overall_score', 0)
        # Convert 0-1 scale to percentage
        row_data['validation_confidence_score'] = round(validation_score * 100, 1)
    elif validation_data and 'validation_report' in validation_data:
        # Fallback to old validation structure
        overall = validation_data['validation_report'].get('overall_assessment', {})
        validation_score = overall.get('confidence_score', 0)
        # Convert 0-1 scale to percentage
        row_data['validation_confidence_score'] = round(validation_score * 100, 1)
    else:
        row_data['validation_confidence_score'] = 0
    
    # Extract average confidence from citation and normalize to percentage (0-100)
    if citation_data and 'citations' in citation_data and 'metadata' in citation_data['citations']:
        metadata = citation_data['citations']['metadata']
        citation_score = metadata.get('average_confidence', 0)
        # Convert 0-1 scale to percentage
        row_data['citation_average_confidence'] = round(citation_score * 100, 1)
    else:
        row_data['citation_average_confidence'] = 0
    
    # Remove image_quality_score column as requested
    # if quality_data and 'overall_assessment' in quality_data:
    #     overall = quality_data['overall_assessment']
    #     row_data['image_quality_score'] = round(overall.get('score', 0), 1)
    # else:
    #     row_data['image_quality_score'] = 0
    
    # Extract LLM quality data from image_quality_assessment in results or separate file
    llm_quality_source = None
    if results_data and 'image_quality_assessment' in results_data:
        llm_quality_source = results_data['image_quality_assessment']
    elif llm_quality_data:
        llm_quality_source = llm_quality_data

    if llm_quality_source:
        # Handle both quality_score (0-100) and overall_quality_score (0-10) formats
        quality_score = llm_quality_source.get('quality_score', 0)
        overall_quality_score = llm_quality_source.get('overall_quality_score', 0)

        if quality_score > 0:
            row_data['llm_image_quality_score'] = round(quality_score, 1)
        elif overall_quality_score > 0:
            # Convert 0-10 scale to percentage
            row_data['llm_image_quality_score'] = round(overall_quality_score * 10, 1)
        else:
            row_data['llm_image_quality_score'] = 0

        row_data['llm_suitable_for_extraction'] = llm_quality_source.get('suitable_for_extraction', False)
    else:
        row_data['llm_image_quality_score'] = 0
        row_data['llm_suitable_for_extraction'] = False
    
    # Calculate overall score with specified weights
    # Weights: validation 60%, llm_image_quality 25%, citation 15% (removed image_quality)
    # All scores are now in percentage format, so normalize to 0-1 for calculation
    validation_score = row_data['validation_confidence_score'] / 100
    citation_score = row_data['citation_average_confidence'] / 100
    llm_score_normalized = row_data['llm_image_quality_score'] / 100
    
    # Apply weights (redistributed after removing image_quality)
    weights = {
        'validation': 0.60,
        'llm_image_quality': 0.25,
        'citation': 0.15
    }
    
    # Calculate weighted overall score (0-1 scale)
    overall_score = (
        validation_score * weights['validation'] +
        llm_score_normalized * weights['llm_image_quality'] +
        citation_score * weights['citation']
    )
    
    # Convert to percentage and round
    row_data['overall_score'] = round(overall_score * 100, 1)
    
    return row_data

def create_timing_summary_worksheet(wb, files):
    """Create a dedicated timing summary worksheet"""
    ws = wb.create_sheet(title="Timing Summary")

    all_timing_data = []

    for filename in files:
        print(f"Processing timing for {filename}...")

        # Load results data
        results_data = (load_json_file(f'results/{filename}_result.json') or
                       load_json_file(f'results/{filename}.json'))

        # Load LLM validation data
        llm_validation_data = None
        validation_files = [
            f'validation_results/{filename}_llm_validation.json',
            f'validation_results/{filename}_validation_*.json'
        ]
        for pattern in validation_files:
            if '*' in pattern:
                # Handle wildcard pattern for timestamped files
                import glob
                matching_files = glob.glob(pattern.replace('*', '*'))
                if matching_files:
                    # Use the most recent file
                    llm_validation_data = load_json_file(sorted(matching_files)[-1])
                    break
            else:
                llm_validation_data = load_json_file(pattern)
                if llm_validation_data:
                    break

        if results_data and 'timing' in results_data:
            timing = results_data['timing']

            # Add total processing time
            total_time = timing.get('total_processing_time_seconds', 'N/A')
            all_timing_data.append({
                'file_name': filename,
                'timing_category': 'Total Processing',
                'phase_or_metric': 'Total Processing Time',
                'duration_seconds': total_time,
                'model_used': '',
                'execution_mode': '',
                'start_time': '',
                'end_time': '',
                'details': 'Complete end-to-end processing time'
            })

            # Add agent performance details
            if 'agent_performance' in timing:
                agent_performance = timing['agent_performance']
                for phase, details in agent_performance.items():
                    if isinstance(details, dict):
                        all_timing_data.append({
                            'file_name': filename,
                            'timing_category': 'Agent Performance',
                            'phase_or_metric': phase.replace('_', ' ').title(),
                            'duration_seconds': details.get('duration_seconds', 'N/A'),
                            'model_used': details.get('model_used', details.get('document_reader_used', '')),
                            'execution_mode': details.get('execution_mode', ''),
                            'start_time': details.get('start_time', ''),
                            'end_time': details.get('end_time', ''),
                            'details': ''
                        })


    # Create DataFrame and add to worksheet
    if all_timing_data:
        df = pd.DataFrame(all_timing_data)
        add_dataframe_to_worksheet(ws, df, 1, "TIMING SUMMARY FOR ALL FILES")

        # Color the total processing time rows to make them stand out
        from openpyxl.styles import PatternFill
        blue_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")

        # Find and color total processing time rows (starting from row 3, after header and title)
        for row_num in range(3, ws.max_row + 1):
            timing_category_cell = ws.cell(row=row_num, column=2)  # Column B contains timing_category
            if timing_category_cell.value == "Total Processing":
                # Color the entire row
                for col_num in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col_num).fill = blue_fill

        # Auto-adjust column widths
        auto_adjust_column_widths(ws)

        print(f"Added timing summary worksheet with {len(all_timing_data)} timing records")
    else:
        # Add empty row if no data
        ws.cell(row=1, column=1, value="No timing data available")
        print("No timing data found")

def create_business_summary_worksheet(wb, file_list):
    """Create business summary worksheet as second tab"""
    print("Creating business summary worksheet...")
    
    # Extract data for all files
    report_data = []
    for file_base in file_list:
        try:
            row_data = extract_business_data(file_base)
            report_data.append(row_data)
        except Exception as e:
            print(f"Error processing {file_base} for business summary: {e}")
    
    # Create DataFrame
    df = pd.DataFrame(report_data)
    
    # Reorder columns for better readability
    column_order = [
        'filename',
        'overall_score',
        'document_location',
        'isExpense',
        'issue_count',
        'validation_confidence_score',
        'citation_average_confidence',
        'llm_image_quality_score',
        'llm_suitable_for_extraction'
    ]
    
    # Only include columns that exist
    final_columns = [col for col in column_order if col in df.columns]
    df = df[final_columns]
    
    # Sort by filename for consistency
    df = df.sort_values('filename').reset_index(drop=True)
    
    # Create business summary worksheet
    ws = wb.create_sheet(title="Business Summary", index=0)
    current_row = 1
    
    # Add business summary data to worksheet
    current_row = add_dataframe_to_worksheet(ws, df, current_row, "BUSINESS SUMMARY REPORT")
    
    # Auto-adjust column widths
    auto_adjust_column_widths(ws)
    
    print(f"Added business summary worksheet with {len(df)} files")

def create_summary_worksheet(wb, file_list):
    """Create summary worksheet as first tab"""
    print("Creating summary worksheet...")
    
    # Initialize data list
    report_data = []
    
    for file_base in file_list:
        # Initialize row data
        row_data = {'file': file_base}
        
        # Load data from current project structure
        results_data = load_json_file(f'results/{file_base}_result.json')

        # Try different validation file patterns (keeping fallback for compatibility)
        validation_data = (load_json_file(f'validation_results/{file_base}_compliance_validation.json') or
                          load_json_file(f'validation_results/{file_base}_validation.json'))

        # Try different quality file patterns (keeping fallback for compatibility)
        quality_data = (load_json_file(f'quality_reports/{file_base}_page1_quality.json') or
                       load_json_file(f'quality_reports/{file_base}_quality.json'))



        # Try different LLM quality file patterns (keeping fallback for compatibility)
        llm_quality_data = (load_json_file(f'llm_quality_reports/llm_quality_{file_base}_pdf.json') or
                           load_json_file(f'llm_quality_reports/llm_quality_{file_base}.json'))
        
        # Extract data from each source
        row_data.update(extract_results_data_summary(results_data))
        row_data.update(extract_validation_data_summary(validation_data))
        row_data.update(extract_quality_data_summary(quality_data))
        row_data.update(extract_llm_quality_data_summary(llm_quality_data, results_data))
        
        # Add dataset info from classification if available
        if results_data and 'classification' in results_data:
            classification = results_data['classification']
            row_data['country'] = classification.get('document_location', '')
            row_data['icp'] = ''  # Not available in current structure
        
        report_data.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(report_data)
    
    # Define column order with QA columns interspersed
    columns_order = ['file']
    data_columns = [
        'issues_count',
        'is_expense',
        'language',
        'language_confidence',
        'classification_confidence',
        'confidence_score',
        'is_reliable',
        'hallucination_detection',
        'compliance_accuracy',
        'factual_grounding',
        'knowledge_base_adherence',
        'quality_score',
        'quality_level',
        'quality_passed',
        'resolution_score',
        'blur_score',
        'glare_score',
        'completeness_score',
        'damage_score',
        'llm_suitable_for_extraction',
        'llm_model_used',
        'llm_blur_detected',
        'llm_blur_severity',
        'llm_glare_detected',
        'llm_glare_severity',
        'llm_cut_off_detected',
        'llm_missing_sections',
        'country',
        'icp'
    ]
    
    # Add each data column followed by QA column
    for col in data_columns:
        if col in df.columns:
            columns_order.append(col)
            columns_order.append('QA')
    
    # Add QA columns to dataframe
    for col in data_columns:
        if col in df.columns:
            df[f'{col}_QA'] = ''
    
    # Reorder columns (only include columns that exist)
    final_columns = []
    for col in columns_order:
        if col == 'QA':
            # Find the previous data column and add its QA column
            prev_col = final_columns[-1] if final_columns else None
            if prev_col and prev_col != 'file':
                qa_col = f'{prev_col}_QA'
                if qa_col in df.columns:
                    final_columns.append(qa_col)
        elif col in df.columns:
            final_columns.append(col)
    
    df = df[final_columns]
    
    # Fill NaN values with appropriate defaults
    df = df.fillna({
        'issues_count': 0,
        'is_expense': False,
        'language': '',
        'language_confidence': 0,
        'classification_confidence': 0,
        'confidence_score': 0,
        'is_reliable': False,
        'hallucination_detection': 0,
        'compliance_accuracy': 0,
        'factual_grounding': 0,
        'knowledge_base_adherence': 0,
        'quality_score': 0,
        'quality_level': '',
        'quality_passed': False,
        'resolution_score': 0,
        'blur_score': 0,
        'glare_score': 0,
        'completeness_score': 0,
        'damage_score': 0,
        'llm_suitable_for_extraction': False,
        'llm_model_used': '',
        'llm_blur_detected': False,
        'llm_blur_severity': '',
        'llm_glare_detected': False,
        'llm_glare_severity': '',
        'llm_cut_off_detected': False,
        'llm_missing_sections': False,
        'country': '',
        'icp': ''
    })
    
    # Create summary worksheet
    ws = wb.create_sheet(title="Summary", index=1)
    current_row = 1
    
    # Add summary data to worksheet
    current_row = add_dataframe_to_worksheet(ws, df, current_row, "EXPENSE ANALYSIS SUMMARY REPORT")
    
    # Auto-adjust column widths
    auto_adjust_column_widths(ws)
    
    print(f"Added summary worksheet with {len(df)} files")

def generate_multitab_report():
    """Generate multi-tab Excel report with one tab per file"""
    file_list = get_file_list()
    print(f"Found {len(file_list)} files to process: {file_list}")
    
    # Create workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create business summary worksheet first
    create_business_summary_worksheet(wb, file_list)

    # Create summary worksheet second
    create_summary_worksheet(wb, file_list)

    # Create timing summary worksheet third
    create_timing_summary_worksheet(wb, file_list)
    
    # Create worksheets for each file
    for filename in file_list:
        try:
            create_worksheet_for_file(wb, filename)
        except Exception as e:
            print(f"Error processing {filename}: {e}")
    
    # Save the workbook
    output_file = 'consolidated_expense_reports.xlsx'
    wb.save(output_file)
    
    print(f"\nSuccessfully generated multi-tab report: {output_file}")
    print(f"Report contains {len(wb.worksheets)} worksheets:")
    for ws in wb.worksheets:
        print(f"  - {ws.title}")
    
    return output_file

if __name__ == "__main__":
    generate_multitab_report()
